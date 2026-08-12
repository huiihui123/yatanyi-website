"""
雅檀怡家私城 - 后端服务 (Flask)
================================
功能：
  1. 接收并保存客户免费设计咨询表单数据
  2. 新闻订阅管理
  3. 管理后台 API（数据看板 + 预约列表 + 订阅管理 + 状态更新）
  4. CORS 支持，允许前端跨域请求
  5. SQLite 轻量数据库，零配置即可运行

启动方式：python app.py
默认端口：5000
管理后台：http://localhost:5000/admin
"""

from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS          # 跨域支持
from datetime import datetime, timedelta
import sqlite3, os, json, hashlib, uuid

# ============================================================
# Flask 应用初始化
# ============================================================
app = Flask(__name__, static_folder=None)  # 不使用Flask默认static
CORS(app)  # 允许前端跨域请求

# 项目根目录（backend的上一级）
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BACKEND_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BACKEND_DIR, 'yatanyi.db')


# ============================================================
# 数据库工具函数
# ============================================================

def get_db():
    """获取数据库连接（自动创建表）"""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row  # 查询结果支持字典访问
    return conn


def init_db():
    """初始化数据库表结构"""
    conn = get_db()
    cursor = conn.cursor()

    # ---- 客户咨询预约表 ----
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS consultations (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            name          TEXT    NOT NULL,
            phone         TEXT    NOT NULL,
            email         TEXT,
            consult_type  TEXT,
            message       TEXT,
            status        TEXT    DEFAULT 'pending',
            created_at    TEXT    NOT NULL,
            ip_address    TEXT
        )
    ''')

    # ---- 新闻订阅表 ----
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS subscriptions (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            email         TEXT    NOT NULL UNIQUE,
            created_at    TEXT    NOT NULL,
            ip_address    TEXT
        )
    ''')

    # ---- 产品表（后续可扩展） ----
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS products (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            name          TEXT    NOT NULL,
            category      TEXT    NOT NULL,
            price         REAL    NOT NULL,
            original_price REAL,
            description   TEXT,
            image_url     TEXT,
            rating        REAL    DEFAULT 4.5,
            review_count  INTEGER DEFAULT 0,
            is_hot        INTEGER DEFAULT 0,
            is_new        INTEGER DEFAULT 0,
            created_at    TEXT    NOT NULL
        )
    ''')

    # ---- 用户表 ----
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            username      TEXT    NOT NULL UNIQUE,
            password      TEXT    NOT NULL,
            phone         TEXT,
            email         TEXT,
            avatar        TEXT    DEFAULT '',
            created_at    TEXT    NOT NULL,
            last_login    TEXT
        )
    ''')

    # ---- 购物车表 ----
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS cart_items (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id       INTEGER NOT NULL,
            product_name  TEXT    NOT NULL,
            product_price REAL    NOT NULL,
            quantity      INTEGER DEFAULT 1,
            image_url     TEXT,
            added_at      TEXT    NOT NULL
        )
    ''')

    # ---- 订单表 ----
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS orders (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id       INTEGER NOT NULL,
            username      TEXT    NOT NULL,
            items_json    TEXT    NOT NULL,
            total_amount  REAL    NOT NULL,
            status        TEXT    DEFAULT 'pending',
            xlsx_path     TEXT,
            created_at    TEXT    NOT NULL
        )
    ''')

    conn.commit()
    conn.close()
    print("[数据库] 初始化完成 ✓")


# ============================================================
# 静态文件服务 — 让管理后台能访问 CSS/JS
# ============================================================

@app.route('/<path:filename>')
def serve_static(filename):
    """提供项目根目录下的静态文件（CSS/JS等）"""
    # 安全检查：只允许特定扩展名
    safe_exts = ('.css', '.js', '.html', '.jpg', '.jpeg', '.png', '.gif', '.svg', '.ico', '.woff', '.woff2', '.ttf')
    if not any(filename.lower().endswith(ext) for ext in safe_exts):
        return jsonify({'error': 'Forbidden'}), 403
    return send_from_directory(PROJECT_ROOT, filename)


# ============================================================
# 管理后台页面
# ============================================================

@app.route('/admin')
@app.route('/admin/')
def admin_panel():
    """管理后台主页"""
    admin_html = os.path.join(BACKEND_DIR, 'admin.html')
    if os.path.exists(admin_html):
        return send_from_directory(BACKEND_DIR, 'admin.html')
    return '<h1>管理后台文件未找到</h1>', 404


# ============================================================
# API 路由 — 前端调用的接口
# ============================================================

@app.route('/api/health', methods=['GET'])
def health_check():
    """健康检查"""
    return jsonify({
        'status': 'ok',
        'service': '雅檀怡家私城后端',
        'timestamp': datetime.now().isoformat()
    })


@app.route('/api/shutdown', methods=['POST'])
def shutdown_server():
    """★ 优雅关闭 — 返回OK后0.5秒进程退出"""
    import os, threading
    def _quit():
        import time; time.sleep(0.5)
        os._exit(0)
    threading.Thread(target=_quit, daemon=True).start()
    return jsonify({'success': True, 'message': '服务正在关闭...'})


@app.route('/api/consult', methods=['POST'])
def submit_consult():
    """
    客户提交免费设计咨询
    前端 JSON: { name, phone, email?, consult_type?, message? }
    """
    try:
        data = request.get_json()

        name = (data.get('name') or '').strip()
        phone = (data.get('phone') or '').strip()

        if not name or not phone:
            return jsonify({'success': False, 'message': '姓名和电话不能为空'}), 400

        if len(name) > 50:
            return jsonify({'success': False, 'message': '姓名过长'}), 400

        if not phone.isdigit() or len(phone) != 11:
            return jsonify({'success': False, 'message': '手机号格式不正确'}), 400

        email = (data.get('email') or '').strip()
        consult_type = (data.get('consult_type') or '').strip()
        message = (data.get('message') or '').strip()
        ip = request.remote_addr or ''

        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO consultations (name, phone, email, consult_type, message, created_at, ip_address)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (name, phone, email, consult_type, message,
              datetime.now().isoformat(), ip))
        conn.commit()
        new_id = cursor.lastrowid
        conn.close()

        print(f"[咨询] 新记录 #{new_id}: {name} {phone} ({consult_type})")

        return jsonify({
            'success': True,
            'message': '提交成功！我们的设计师将在24小时内与您联系。',
            'id': new_id
        })

    except Exception as e:
        print(f"[错误] 提交咨询失败: {e}")
        return jsonify({'success': False, 'message': '服务器错误，请稍后重试'}), 500


@app.route('/api/subscribe', methods=['POST'])
def subscribe_newsletter():
    """新闻/资讯订阅"""
    try:
        data = request.get_json()
        email = (data.get('email') or '').strip()

        if not email or '@' not in email:
            return jsonify({'success': False, 'message': '请输入有效的邮箱地址'}), 400

        ip = request.remote_addr or ''

        conn = get_db()
        cursor = conn.cursor()

        cursor.execute('SELECT id FROM subscriptions WHERE email = ?', (email,))
        if cursor.fetchone():
            conn.close()
            return jsonify({'success': True, 'message': '您已订阅，无需重复订阅'})

        cursor.execute(
            'INSERT INTO subscriptions (email, created_at, ip_address) VALUES (?, ?, ?)',
            (email, datetime.now().isoformat(), ip))
        conn.commit()
        conn.close()

        print(f"[订阅] {email}")
        return jsonify({'success': True, 'message': '订阅成功！感谢您的关注。'})

    except Exception as e:
        print(f"[错误] 订阅失败: {e}")
        return jsonify({'success': False, 'message': '服务器错误'}), 500


# ============================================================
# 管理端 API — 后台数据管理
# ============================================================

@app.route('/api/stats', methods=['GET'])
def get_stats():
    """
    数据看板统计
    返回：总咨询数、待处理数、今日新增、总订阅数
    """
    conn = get_db()
    cursor = conn.cursor()

    # 总咨询数
    cursor.execute('SELECT COUNT(*) FROM consultations')
    total_consult = cursor.fetchone()[0]

    # 待处理数
    cursor.execute("SELECT COUNT(*) FROM consultations WHERE status='pending'")
    pending = cursor.fetchone()[0]

    # 今日新增
    today = datetime.now().strftime('%Y-%m-%d')
    cursor.execute("SELECT COUNT(*) FROM consultations WHERE created_at LIKE ?", (today + '%',))
    today_new = cursor.fetchone()[0]

    # 总订阅数
    cursor.execute('SELECT COUNT(*) FROM subscriptions')
    total_subs = cursor.fetchone()[0]

    # 今日订阅
    cursor.execute("SELECT COUNT(*) FROM subscriptions WHERE created_at LIKE ?", (today + '%',))
    today_subs = cursor.fetchone()[0]

    # 已完成数
    cursor.execute("SELECT COUNT(*) FROM consultations WHERE status='completed'")
    completed = cursor.fetchone()[0]

    conn.close()

    return jsonify({
        'success': True,
        'data': {
            'total_consult': total_consult,
            'pending': pending,
            'today_new': today_new,
            'completed': completed,
            'total_subs': total_subs,
            'today_subs': today_subs
        }
    })


@app.route('/api/admin/consultations', methods=['GET'])
def list_consultations():
    """
    查看所有咨询记录
    支持参数：?status=pending&page=1&limit=20
    """
    status_filter = request.args.get('status', '')
    page = int(request.args.get('page', 1))
    limit = int(request.args.get('limit', 50))
    offset = (page - 1) * limit

    conn = get_db()
    cursor = conn.cursor()

    # 条件查询
    if status_filter:
        cursor.execute('SELECT COUNT(*) FROM consultations WHERE status = ?', (status_filter,))
        total = cursor.fetchone()[0]
        cursor.execute(
            'SELECT * FROM consultations WHERE status = ? ORDER BY created_at DESC LIMIT ? OFFSET ?',
            (status_filter, limit, offset))
    else:
        cursor.execute('SELECT COUNT(*) FROM consultations')
        total = cursor.fetchone()[0]
        cursor.execute(
            'SELECT * FROM consultations ORDER BY created_at DESC LIMIT ? OFFSET ?',
            (limit, offset))

    rows = [dict(row) for row in cursor.fetchall()]
    conn.close()

    return jsonify({
        'success': True,
        'data': rows,
        'total': total,
        'page': page,
        'limit': limit
    })


@app.route('/api/admin/subscriptions', methods=['GET'])
def list_subscriptions():
    """查看所有订阅记录"""
    page = int(request.args.get('page', 1))
    limit = int(request.args.get('limit', 50))
    offset = (page - 1) * limit

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT COUNT(*) FROM subscriptions')
    total = cursor.fetchone()[0]
    cursor.execute('SELECT * FROM subscriptions ORDER BY created_at DESC LIMIT ? OFFSET ?', (limit, offset))
    rows = [dict(row) for row in cursor.fetchall()]
    conn.close()

    return jsonify({
        'success': True,
        'data': rows,
        'total': total,
        'page': page,
        'limit': limit
    })


@app.route('/api/admin/consultations/<int:cid>/status', methods=['PUT'])
def update_consultation_status(cid):
    """更新咨询状态：pending → contacted → completed"""
    data = request.get_json()
    new_status = (data.get('status') or '').strip()

    if new_status not in ('pending', 'contacted', 'completed'):
        return jsonify({'success': False, 'message': '无效状态'}), 400

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('UPDATE consultations SET status = ? WHERE id = ?', (new_status, cid))
    conn.commit()
    affected = cursor.rowcount
    conn.close()

    if affected == 0:
        return jsonify({'success': False, 'message': '记录不存在'}), 404

    return jsonify({'success': True, 'message': f'状态已更新为 {new_status}'})


@app.route('/api/admin/consultations/<int:cid>', methods=['DELETE'])
def delete_consultation(cid):
    """删除咨询记录"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM consultations WHERE id = ?', (cid,))
    conn.commit()
    affected = cursor.rowcount
    conn.close()

    if affected == 0:
        return jsonify({'success': False, 'message': '记录不存在'}), 404

    return jsonify({'success': True, 'message': '记录已删除'})


@app.route('/api/admin/subscriptions/<int:sid>', methods=['DELETE'])
def delete_subscription(sid):
    """删除订阅记录"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM subscriptions WHERE id = ?', (sid,))
    conn.commit()
    affected = cursor.rowcount
    conn.close()

    if affected == 0:
        return jsonify({'success': False, 'message': '记录不存在'}), 404

    return jsonify({'success': True, 'message': '订阅已删除'})


# ============================================================
# ★★★ 用户认证 API ★★★
# ============================================================

def _hash_pw(pw):
    """简单密码哈希（生产环境应用 bcrypt）"""
    return hashlib.sha256(pw.encode()).hexdigest()

@app.route('/api/auth/register', methods=['POST'])
def register():
    """用户注册"""
    data = request.get_json()
    username = (data.get('username') or '').strip()
    password = (data.get('password') or '').strip()
    phone = (data.get('phone') or '').strip()
    if not username or len(username) < 2:
        return jsonify({'success': False, 'message': '用户名至少2个字符'}), 400
    if not password or len(password) < 6:
        return jsonify({'success': False, 'message': '密码至少6位'}), 400
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT id FROM users WHERE username=?', (username,))
    if c.fetchone():
        conn.close()
        return jsonify({'success': False, 'message': '用户名已存在'}), 400
    token = uuid.uuid4().hex
    c.execute('INSERT INTO users (username,password,phone,created_at,last_login) VALUES (?,?,?,?,?)',
              (username, _hash_pw(password), phone, datetime.now().isoformat(), datetime.now().isoformat()))
    user_id = c.lastrowid
    conn.commit(); conn.close()
    # 创建购物车token（用token做客户端身份标识）
    return jsonify({'success': True, 'message': '注册成功！', 'user': {'id': user_id, 'username': username, 'token': token}})

@app.route('/api/auth/login', methods=['POST'])
def login():
    """用户登录"""
    data = request.get_json()
    username = (data.get('username') or '').strip()
    password = (data.get('password') or '').strip()
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT * FROM users WHERE username=?', (username,))
    row = c.fetchone()
    if not row or row['password'] != _hash_pw(password):
        conn.close()
        return jsonify({'success': False, 'message': '用户名或密码错误'}), 401
    c.execute('UPDATE users SET last_login=? WHERE id=?', (datetime.now().isoformat(), row['id']))
    conn.commit(); conn.close()
    return jsonify({'success': True, 'message': '登录成功！',
        'user': {'id': row['id'], 'username': row['username'], 'phone': row['phone'],
                 'email': row['email'], 'nickname': row['nickname'] or row['username'],
                 'gender': row['gender'], 'age': row['age'], 'region': row['region'],
                 'signature': row['signature'], 'avatar': row['avatar']}})


# ============================================================
# ★★★ 忘记密码 API（验证码方式） ★★★
# ============================================================

import random as _random

@app.route('/api/auth/forgot-password', methods=['POST'])
def forgot_password():
    """发送验证码到手机（模拟：生成6位随机码并返回）"""
    data = request.get_json()
    phone = (data.get('phone') or '').strip()
    if not phone or len(phone) != 11 or not phone.isdigit():
        return jsonify({'success': False, 'message': '请输入正确的11位手机号'}), 400
    conn = get_db(); c = conn.cursor()
    c.execute('SELECT id,username FROM users WHERE phone=?', (phone,))
    user = c.fetchone()
    if not user:
        conn.close()
        return jsonify({'success': False, 'message': '该手机号未注册'}), 404
    # 生成6位验证码，有效期5分钟
    code = str(_random.randint(100000, 999999))
    expires = (datetime.now() + timedelta(minutes=5)).isoformat()
    c.execute('INSERT INTO verification_codes (phone,code,purpose,expires_at,created_at) VALUES (?,?,?,?,?)',
              (phone, code, 'reset_password', expires, datetime.now().isoformat()))
    conn.commit(); conn.close()
    print(f'[验证码] {phone} → {code}（有效至 {expires}）')
    return jsonify({'success': True, 'message': f'验证码已发送到 {phone}',
                    'code': code,  # TOOD: 生产环境删除此行，通过短信发送
                    'user_id': user['id'], 'username': user['username']})

@app.route('/api/auth/verify-code', methods=['POST'])
def verify_code():
    """验证验证码是否正确"""
    data = request.get_json()
    phone = (data.get('phone') or '').strip()
    code = (data.get('code') or '').strip()
    if not phone or not code:
        return jsonify({'success': False, 'message': '参数不完整'}), 400
    conn = get_db(); c = conn.cursor()
    c.execute('''SELECT * FROM verification_codes WHERE phone=? AND code=? AND purpose='reset_password'
                 AND used=0 AND expires_at > ? ORDER BY created_at DESC LIMIT 1''',
              (phone, code, datetime.now().isoformat()))
    row = c.fetchone()
    if not row:
        conn.close()
        return jsonify({'success': False, 'message': '验证码错误或已过期'}), 400
    c.execute('UPDATE verification_codes SET used=1 WHERE id=?', (row['id'],))
    conn.commit(); conn.close()
    return jsonify({'success': True, 'message': '验证通过', 'phone': phone})

@app.route('/api/auth/reset-password', methods=['POST'])
def reset_password():
    """重置密码"""
    data = request.get_json()
    phone = (data.get('phone') or '').strip()
    code = (data.get('code') or '').strip()
    password = (data.get('password') or '').strip()
    if not phone or not code or not password or len(password) < 6:
        return jsonify({'success': False, 'message': '密码至少6位'}), 400
    conn = get_db(); c = conn.cursor()
    # 再次验证验证码
    c.execute('''SELECT * FROM verification_codes WHERE phone=? AND code=? AND purpose='reset_password'
                 AND used=1 AND expires_at > ? ORDER BY created_at DESC LIMIT 1''',
              (phone, code, datetime.now().isoformat()))
    if not c.fetchone():
        conn.close()
        return jsonify({'success': False, 'message': '验证码无效'}), 400
    c.execute('UPDATE users SET password=? WHERE phone=?', (_hash_pw(password), phone))
    conn.commit(); conn.close()
    return jsonify({'success': True, 'message': '密码重置成功！请重新登录。'})


# ============================================================
# ★★★ 个人资料 API ★★★
# ============================================================

@app.route('/api/user/profile', methods=['GET'])
def get_profile():
    """获取用户个人资料"""
    user_id = request.args.get('user_id')
    if not user_id:
        return jsonify({'success': False, 'message': '请先登录'}), 401
    conn = get_db(); c = conn.cursor()
    c.execute('SELECT id,username,phone,email,nickname,gender,age,region,signature,avatar,created_at,last_login FROM users WHERE id=?', (user_id,))
    row = c.fetchone()
    conn.close()
    if not row:
        return jsonify({'success': False, 'message': '用户不存在'}), 404
    return jsonify({'success': True, 'user': dict(row)})

@app.route('/api/user/profile', methods=['PUT'])
def update_profile():
    """更新个人资料"""
    data = request.get_json()
    user_id = data.get('user_id')
    if not user_id:
        return jsonify({'success': False, 'message': '请先登录'}), 401
    fields = ['nickname','gender','age','region','signature','avatar','phone','email']
    updates = {}
    for f in fields:
        if f in data:
            updates[f] = str(data[f]).strip()
    if not updates:
        return jsonify({'success': False, 'message': '没有要更新的字段'}), 400
    conn = get_db(); c = conn.cursor()
    set_clause = ', '.join(f'{k}=?' for k in updates)
    vals = list(updates.values()) + [user_id]
    c.execute(f'UPDATE users SET {set_clause} WHERE id=?', vals)
    conn.commit(); conn.close()
    return jsonify({'success': True, 'message': '资料更新成功！'})

@app.route('/api/user/account', methods=['DELETE'])
def delete_account():
    """注销账户（清除个人信息，保留订单记录）"""
    user_id = request.args.get('user_id')
    if not user_id:
        return jsonify({'success': False, 'message': '请先登录'}), 401
    conn = get_db(); c = conn.cursor()
    c.execute('SELECT username FROM users WHERE id=?', (user_id,))
    row = c.fetchone()
    if not row:
        conn.close()
        return jsonify({'success': False, 'message': '用户不存在'}), 404
    # 清除个人信息（保留ID和订单关联）
    c.execute('''UPDATE users SET username=?, password='', phone='', email='', 
                 nickname='已注销', gender='', age=0, region='', signature='该用户已注销', avatar='' 
                 WHERE id=?''', ('deleted_' + str(user_id), user_id))
    c.execute('DELETE FROM cart_items WHERE user_id=?', (user_id,))
    conn.commit(); conn.close()
    return jsonify({'success': True, 'message': '账户已注销。感谢您的使用。'})


# ============================================================
# ★★★ 购物车 API ★★★
# ============================================================

@app.route('/api/cart', methods=['GET'])
def get_cart():
    """获取用户购物车"""
    user_id = request.args.get('user_id')
    if not user_id:
        return jsonify({'success': False, 'message': '请先登录'}), 401
    conn = get_db(); c = conn.cursor()
    c.execute('SELECT * FROM cart_items WHERE user_id=? ORDER BY added_at DESC', (user_id,))
    items = [dict(r) for r in c.fetchall()]
    total = sum(i['product_price'] * i['quantity'] for i in items)
    conn.close()
    return jsonify({'success': True, 'items': items, 'count': len(items), 'total': total})

@app.route('/api/cart', methods=['POST'])
def add_to_cart():
    """添加商品到购物车"""
    data = request.get_json()
    user_id = data.get('user_id')
    product_name = data.get('product_name')
    product_price = data.get('product_price')
    quantity = int(data.get('quantity', 1))
    image_url = data.get('image_url', '')
    if not user_id or not product_name:
        return jsonify({'success': False, 'message': '参数不完整'}), 400
    conn = get_db(); c = conn.cursor()
    # 检查是否已有同产品
    c.execute('SELECT id,quantity FROM cart_items WHERE user_id=? AND product_name=?',
              (user_id, product_name))
    existing = c.fetchone()
    if existing:
        c.execute('UPDATE cart_items SET quantity=quantity+?, added_at=? WHERE id=?',
                  (quantity, datetime.now().isoformat(), existing['id']))
    else:
        c.execute('INSERT INTO cart_items (user_id,product_name,product_price,quantity,image_url,added_at) VALUES (?,?,?,?,?,?)',
                  (user_id, product_name, product_price, quantity, image_url, datetime.now().isoformat()))
    conn.commit(); conn.close()
    return jsonify({'success': True, 'message': '已加入购物车'})

@app.route('/api/cart/<int:cid>', methods=['PUT'])
def update_cart_item(cid):
    """更新购物车数量"""
    data = request.get_json()
    qty = int(data.get('quantity', 1))
    if qty < 1:
        return jsonify({'success': False, 'message': '数量至少为1'}), 400
    conn = get_db(); c = conn.cursor()
    c.execute('UPDATE cart_items SET quantity=? WHERE id=?', (qty, cid))
    conn.commit(); conn.close()
    return jsonify({'success': True})

@app.route('/api/cart/<int:cid>', methods=['DELETE'])
def delete_cart_item(cid):
    """删除购物车项"""
    conn = get_db(); c = conn.cursor()
    c.execute('DELETE FROM cart_items WHERE id=?', (cid,))
    conn.commit(); conn.close()
    return jsonify({'success': True, 'message': '已移除'})


# ============================================================
# ★★★ 订单 API（生成xlsx） ★★★
# ============================================================

ORDERS_DIR = os.path.join(BACKEND_DIR, '订单文件')
os.makedirs(ORDERS_DIR, exist_ok=True)

def _generate_order_xlsx(order_id, username, items, total):
    """生成订单xlsx文件（只有客服/管理能看）"""
    try:
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active
        ws.title = f'订单_{order_id}'
        ws['A1'] = '订单编号'; ws['B1'] = order_id
        ws['A2'] = '客户姓名'; ws['B2'] = username
        ws['A3'] = '下单时间'; ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A5'] = '产品名称'; ws['B5'] = '单价'; ws['C5'] = '数量'; ws['D5'] = '小计'
        row = 6
        for item in items:
            ws.cell(row, 1, item.get('product_name',''))
            ws.cell(row, 2, item.get('product_price',0))
            ws.cell(row, 3, item.get('quantity',1))
            ws.cell(row, 4, item.get('product_price',0) * item.get('quantity',1))
            row += 1
        ws.cell(row+1, 1, '合计'); ws.cell(row+1, 4, total)
        filename = f'order_{order_id}_{username}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        filepath = os.path.join(ORDERS_DIR, filename)
        wb.save(filepath)
        return filepath
    except Exception as e:
        print(f'[xlsx] 生成失败: {e}')
        return None

@app.route('/api/orders', methods=['POST'])
def create_order():
    """下单：生成订单记录 + xlsx文件"""
    data = request.get_json()
    user_id = data.get('user_id')
    username = data.get('username', '未知用户')
    if not user_id:
        return jsonify({'success': False, 'message': '请先登录'}), 401
    # 获取购物车物品
    conn = get_db(); c = conn.cursor()
    c.execute('SELECT * FROM cart_items WHERE user_id=?', (user_id,))
    cart_items = [dict(r) for r in c.fetchall()]
    if not cart_items:
        conn.close()
        return jsonify({'success': False, 'message': '购物车为空'}), 400
    items_data = [{'product_name': i['product_name'], 'product_price': i['product_price'],
                    'quantity': i['quantity'], 'image_url': i.get('image_url','')} for i in cart_items]
    total = sum(i['product_price'] * i['quantity'] for i in cart_items)
    # 写订单记录
    c.execute('INSERT INTO orders (user_id,username,items_json,total_amount,created_at) VALUES (?,?,?,?,?)',
              (user_id, username, json.dumps(items_data, ensure_ascii=False), total, datetime.now().isoformat()))
    order_id = c.lastrowid
    # 生成xlsx
    xlsx_path = _generate_order_xlsx(order_id, username, items_data, total)
    if xlsx_path:
        c.execute('UPDATE orders SET xlsx_path=? WHERE id=?', (xlsx_path, order_id))
    # 清空购物车
    c.execute('DELETE FROM cart_items WHERE user_id=?', (user_id,))
    conn.commit(); conn.close()
    return jsonify({'success': True, 'message': f'下单成功！共{len(cart_items)}件，¥{total}',
                    'order_id': order_id, 'total': total, 'item_count': len(cart_items)})

@app.route('/api/orders', methods=['GET'])
def get_orders():
    """获取用户订单"""
    user_id = request.args.get('user_id')
    if not user_id:
        return jsonify({'success': False, 'message': '请先登录'}), 401
    conn = get_db(); c = conn.cursor()
    c.execute('SELECT * FROM orders WHERE user_id=? ORDER BY created_at DESC', (user_id,))
    orders = [dict(r) for r in c.fetchall()]
    conn.close()
    return jsonify({'success': True, 'orders': orders})


# ============================================================
# ★★★ 产品搜索 API ★★★
# ============================================================

@app.route('/api/products/search', methods=['GET'])
def search_products():
    """产品搜索 + 排序"""
    q = request.args.get('q', '')
    category = request.args.get('category', '')
    sort = request.args.get('sort', '')  # price_asc/price_desc/rating
    conn = get_db(); c = conn.cursor()
    sql = 'SELECT * FROM products WHERE 1=1'
    params = []
    if q:
        sql += ' AND name LIKE ?'
        params.append(f'%{q}%')
    if category:
        sql += ' AND category=?'
        params.append(category)
    if sort == 'price_asc':
        sql += ' ORDER BY price ASC'
    elif sort == 'price_desc':
        sql += ' ORDER BY price DESC'
    elif sort == 'rating':
        sql += ' ORDER BY rating DESC'
    else:
        sql += ' ORDER BY id DESC'
    c.execute(sql, params)
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return jsonify({'success': True, 'products': rows, 'total': len(rows)})


# ============================================================
# ★★★ 管理端：用户 & 订单管理 ★★★
# ============================================================

@app.route('/api/admin/users', methods=['GET'])
def admin_users():
    """管理端：查看所有用户"""
    conn = get_db(); c = conn.cursor()
    c.execute('SELECT id,username,phone,email,created_at,last_login FROM users ORDER BY created_at DESC')
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return jsonify({'success': True, 'users': rows, 'total': len(rows)})

@app.route('/api/admin/orders', methods=['GET'])
def admin_orders():
    """管理端：查看所有订单"""
    conn = get_db(); c = conn.cursor()
    c.execute('SELECT * FROM orders ORDER BY created_at DESC')
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return jsonify({'success': True, 'orders': rows, 'total': len(rows)})

@app.route('/api/admin/orders/<int:oid>/xlsx', methods=['GET'])
def download_order_xlsx(oid):
    """下载订单xlsx文件（管理端）"""
    conn = get_db(); c = conn.cursor()
    c.execute('SELECT xlsx_path FROM orders WHERE id=?', (oid,))
    row = c.fetchone()
    conn.close()
    if not row or not row['xlsx_path'] or not os.path.exists(row['xlsx_path']):
        return jsonify({'success': False, 'message': '文件不存在'}), 404
    return send_file(row['xlsx_path'], as_attachment=True, download_name=os.path.basename(row['xlsx_path']))


# ============================================================
# 启动入口
# ============================================================

if __name__ == '__main__':
    init_db()

    print("=" * 60)
    print("  雅檀怡家私城 - 后端服务")
    print(f"  项目目录: {PROJECT_ROOT}")
    print("  前端首页: http://localhost:5000/index.html")
    print("  管理后台: http://localhost:5000/admin")
    print("  健康检查: http://localhost:5000/api/health")
    print("=" * 60)

    # ★ debug=False 避免Flask reloader在Windows下产生孤儿进程
    # ★ SO_REUSEADDR 避免频繁重启时端口TIME_WAIT导致绑定失败
    import socket as _socket
    from werkzeug.serving import BaseWSGIServer as _BWSGI
    _orig_init = _BWSGI.__init__
    def _patched_init(self, *args, **kwargs):
        _orig_init(self, *args, **kwargs)
        self.socket.setsockopt(_socket.SOL_SOCKET, _socket.SO_REUSEADDR, 1)
    _BWSGI.__init__ = _patched_init
    app.run(host='0.0.0.0', port=5000, debug=False, threaded=True)
