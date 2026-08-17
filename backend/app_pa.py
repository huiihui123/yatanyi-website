"""
雅檀怡家私城 - PythonAnywhere 部署版（由 app.py 同步生成，包含全部安全修复）
==========================================================================
★ 本文件与 app.py 保持同一套代码，仅路径适配 PythonAnywhere。
   修改功能请改 app.py 后重新同步本文件（或直接部署 app.py）。
"""
import os
import sys

# PythonAnywhere 环境检测
IS_PA = os.path.exists('/var/run/apache2')

if IS_PA:
    # 自动获取 PythonAnywhere 的 HOME 目录（/home/你的用户名）
    HOME_DIR = os.environ.get('HOME', os.path.expanduser('~'))
    PROJECT_ROOT = os.path.join(HOME_DIR, 'mysite')
    BACKEND_DIR = os.path.join(PROJECT_ROOT, 'backend')
else:
    PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    BACKEND_DIR = os.path.dirname(os.path.abspath(__file__))

DB_PATH = os.path.join(BACKEND_DIR, 'yatanyi.db')

# 确保 images 目录在后端可访问
IMAGES_DIR = os.path.join(PROJECT_ROOT, 'images')
os.makedirs(IMAGES_DIR, exist_ok=True)

from flask import Flask, request, jsonify, send_from_directory, send_file, session
from flask_cors import CORS          # 跨域支持
from datetime import datetime, timedelta
from functools import wraps
import sqlite3, os, json, hashlib, uuid, threading, sys

# ★ BUG修复：中文Windows控制台默认GBK编码，print含✓等Unicode字符会抛
#   UnicodeEncodeError 导致程序启动即崩溃。统一改为UTF-8输出。
if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if sys.stderr and hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

# ============================================================
# Flask 应用初始化
# ============================================================
app = Flask(__name__, static_folder=None)  # 不使用Flask默认static
# ★ 安全加固：secret_key 优先取环境变量，避免源码泄露导致可伪造管理员会话
app.secret_key = os.environ.get('YATANYI_SECRET_KEY', 'yatanyi-admin-secret-key-2026-v1')
# ★ 安全加固：session cookie 只走 HttpOnly + SameSite=Lax（防跨站请求携带）
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    PERMANENT_SESSION_LIFETIME=timedelta(hours=8),
)
# ★ 安全加固：CORS 默认只允许同源；如确需跨域可改 YATANYI_ORIGINS 环境变量（逗号分隔）
_allowed_origins = [o.strip() for o in os.environ.get('YATANYI_ORIGINS', '').split(',') if o.strip()]
if _allowed_origins:
    CORS(app, origins=_allowed_origins, supports_credentials=True)
else:
    CORS(app, resources={r"/api/*": {"origins": "*"}})  # API 无 cookie 跨域需求，保持宽松但不影响 session 安全

# ★ 管理后台密码（优先环境变量；未设置时用默认值，请务必修改！）
ADMIN_PASSWORD = os.environ.get('YATANYI_ADMIN_PASSWORD', '12345hhh')

# 登录失败锁定：同一IP连续失败5次锁定10分钟，防暴力破解
_login_fails = {}   # {ip: [count, lock_until]}
_LOCK_MIN = 10

def _check_login_lock(ip):
    rec = _login_fails.get(ip)
    if rec and rec[1] and datetime.now() < rec[1]:
        # ★ BUG修复：now < rec[1] 时 (now - rec[1]).seconds 会得到 86399-剩余秒 的错误值，
        #   应反过来用 (rec[1] - now).seconds 才是真实剩余秒数
        return (rec[1] - datetime.now()).seconds
    return None

def _record_login_fail(ip):
    now = datetime.now()
    rec = _login_fails.get(ip)
    # ★ BUG修复：原实现 rec[1] 为 None 时 `now >= rec[1]` 抛 TypeError（锁定失效）
    if rec is None:
        rec = [1, None]
    elif rec[1] is not None and now >= rec[1]:
        rec[0] = 1
        rec[1] = None
    else:
        rec[0] += 1
    if rec[0] >= 5:
        rec[1] = now + timedelta(minutes=_LOCK_MIN)
        rec[0] = 0
    _login_fails[ip] = rec

# 登录保护装饰器：未登录一律 401
def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get('admin_auth'):
            return jsonify({'success': False, 'message': '未登录'}), 401
        return f(*args, **kwargs)
    return wrapper

# ★ 安全加固：用户端登录保护——身份一律从服务端 session 取，
#   彻底杜绝「信任前端传 user_id」导致的越权（IDOR）漏洞
def require_login(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        uid = session.get('user_id')
        if not uid:
            return jsonify({'success': False, 'message': '请先登录'}), 401
        return f(*args, **kwargs)
    return wrapper

def current_user_id():
    """返回当前登录用户 id（session 优先），兼容旧版前端传参"""
    return session.get('user_id')

# 项目根目录（backend的上一级）
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BACKEND_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BACKEND_DIR, 'yatanyi.db')

# SQLite 并发：WAL + busy_timeout，避免频繁刷新时 database is locked / 线程堵死
_DB_INIT_LOCK = threading.Lock()
_DB_READY = False

# ★ 线程级连接池：每个工作线程复用一条连接，不再每次请求新建+重设WAL
#   （WAL 是数据库级持久设置，init_db 设一次即可，重复设置会抢写锁导致并发卡死）
_local = threading.local()


@app.after_request
def _perf_headers(resp):
    """分级缓存策略（2026-08-18 优化版）：
       - 图片等二进制资源：强缓存 30 天（产品图文件名带时间戳天然唯一，
         不常变的图没必要频繁重下；长缓存显著提升二次访问速度）
       - CSS/JS 带版本号(?v=xxx)：强缓存 1 年 + immutable —— 内容一变就改版本号
         换新URL，浏览器自动请求新文件；没变的文件继续命中缓存。
         即"及时更新靠换URL、加载加速靠长缓存"，两者兼得
       - CSS/JS 未带版本号：降级为 no-cache 协商缓存（兜底：万一改了文件忘了
         加版本号，也不会让用户拿到旧版卡死）
       - HTML（含 / 首页、/admin 后台）：no-cache 协商缓存 —— 每次向服务器
         校验（配合 ETag/Last-Modified 返回 304，几乎零流量），有更新立即生效
       - API：no-store，保证数据永远最新"""
    path = request.path.lower()
    # ★ BUG修复：404/错误响应不能加缓存头，否则浏览器会把"图片加载失败"缓存 30 天，
    #   修复图片后用户刷新也看不到图。仅在请求成功(2xx)时才允许强缓存。
    if resp.status_code >= 200 and resp.status_code < 300 and path.endswith(('.jpg', '.jpeg', '.png', '.gif', '.svg', '.woff', '.woff2', '.ttf', '.ico')):
        # ★ SWR（Stale-While-Revalidate）并发式缓存：30 天内直接命中缓存秒开；
        #   过期后 1 天内先用旧缓存立即响应，同时后台并发向服务器拉新版本更新缓存——
        #   用户永远不等待，图片更新最多滞后 1 天，两全其美（Chrome 68+/Safari 17+ 支持，
        #   旧浏览器自动忽略 SWR 指令，退化为普通 30 天缓存，无副作用）
        resp.headers['Cache-Control'] = 'public, max-age=2592000, stale-while-revalidate=86400'  # 图片：30天 + SWR 1天
    elif path.startswith('/api/'):
        resp.headers['Cache-Control'] = 'no-store'                            # 接口：不缓存
    elif path.endswith(('.css', '.js')):
        if request.args.get('v'):                                            # 带版本号指纹 → 长缓存
            resp.headers['Cache-Control'] = 'public, max-age=31536000, immutable'  # 1 年
        else:                                                                # 未带版本号 → 协商缓存兜底
            resp.headers['Cache-Control'] = 'no-cache'
    elif path == '/' or path.endswith('.html') or path == '/admin' or path.startswith('/admin/'):
        resp.headers['Cache-Control'] = 'no-cache'                            # 页面：每次校验，未变 304
    return resp


# ============================================================
# 数据库工具函数（请求级连接，必定关闭）
# ============================================================

def _connect_db():
    """创建带超时与 WAL 的 SQLite 连接（WAL 是数据库级设置，连接上不再重复设置）"""
    conn = sqlite3.connect(
        DB_PATH,
        timeout=5,                  # 锁等待最多 5 秒，快速失败而非无限挂起
        check_same_thread=False,    # 配合 Flask threaded=True
    )
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA busy_timeout=5000')     # 等锁最多 5 秒（原来30秒→刷新风暴时线程全挂死）
    conn.execute('PRAGMA synchronous=NORMAL')    # WAL 下性能更好且足够安全
    conn.execute('PRAGMA foreign_keys=ON')
    conn.execute('PRAGMA temp_store=MEMORY')
    return conn


def get_db():
    """获取当前线程的数据库连接（线程级复用，请求结束不关闭，归还原连接）"""
    conn = getattr(_local, 'conn', None)
    if conn is None:
        conn = _connect_db()
        _local.conn = conn
    else:
        # 保险：连接可能因锁异常处于坏状态，探测后重建
        try:
            conn.execute('SELECT 1')
        except Exception:
            try:
                conn.close()
            except Exception:
                pass
            conn = _connect_db()
            _local.conn = conn
    return conn


@app.teardown_appcontext
def close_db(exc=None):
    """请求结束：回滚未提交事务，连接归还线程池复用（不真正关闭，避免频繁开关连接）"""
    conn = getattr(_local, 'conn', None)
    if conn is not None:
        try:
            conn.rollback()   # 清掉残留事务，连接回到干净状态
        except Exception:
            pass


def init_db():
    """初始化数据库表结构（含字段迁移）"""
    global _DB_READY
    with _DB_INIT_LOCK:
        if _DB_READY and os.path.exists(DB_PATH):
            return
        conn = _connect_db()
        # ★ WAL 是数据库级持久设置，只在初始化时设置一次
        conn.execute('PRAGMA journal_mode=WAL')
        cursor = conn.cursor()

        # ---- 客户咨询预约表 ----
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS consultations (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                name          TEXT    NOT NULL,
                phone         TEXT    NOT NULL,
                email         TEXT,
                consult_type  TEXT,
                message       TEXT,
                status        TEXT    DEFAULT 'pending',
                created_at    TEXT    NOT NULL,
                ip_address    TEXT
            )
        ''')

        # ---- 新闻订阅表 ----
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS subscriptions (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                email         TEXT    NOT NULL UNIQUE,
                created_at    TEXT    NOT NULL,
                ip_address    TEXT
            )
        ''')

        # ---- 产品表 ----
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS products (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                name          TEXT    NOT NULL,
                category      TEXT    NOT NULL,
                price         REAL    NOT NULL,
                original_price REAL,
                description   TEXT,
                image_url     TEXT,
                rating        REAL    DEFAULT 4.5,
                review_count  INTEGER DEFAULT 0,
                is_hot        INTEGER DEFAULT 0,
                is_new        INTEGER DEFAULT 0,
                is_active     INTEGER DEFAULT 1,
                created_at    TEXT    NOT NULL
            )
        ''')
        # ★ 老库迁移：products 表补充 is_active（上架状态，1=上架 0=下架）
        try:
            cursor.execute('ALTER TABLE products ADD COLUMN is_active INTEGER DEFAULT 1')
        except sqlite3.OperationalError:
            pass

        # ---- 用户表（含个人资料字段） ----
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                username      TEXT    NOT NULL UNIQUE,
                password      TEXT    NOT NULL,
                phone         TEXT,
                email         TEXT,
                nickname      TEXT    DEFAULT '',
                gender        TEXT    DEFAULT '',
                age           INTEGER DEFAULT 0,
                region        TEXT    DEFAULT '',
                signature     TEXT    DEFAULT '',
                avatar        TEXT    DEFAULT '',
                created_at    TEXT    NOT NULL,
                last_login    TEXT
            )
        ''')
        for col, typ in (
            ('nickname', 'TEXT DEFAULT ""'),
            ('gender', 'TEXT DEFAULT ""'),
            ('age', 'INTEGER DEFAULT 0'),
            ('region', 'TEXT DEFAULT ""'),
            ('signature', 'TEXT DEFAULT ""'),
            ('avatar', 'TEXT DEFAULT ""'),
        ):
            try:
                cursor.execute(f'ALTER TABLE users ADD COLUMN {col} {typ}')
            except sqlite3.OperationalError:
                pass

        # ---- 验证码表 ----
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS verification_codes (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                phone         TEXT    NOT NULL,
                code          TEXT    NOT NULL,
                purpose       TEXT    DEFAULT 'reset_password',
                expires_at    TEXT    NOT NULL,
                used          INTEGER DEFAULT 0,
                created_at    TEXT    NOT NULL
            )
        ''')

        # ---- 购物车表 ----
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS cart_items (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id       INTEGER NOT NULL,
                product_name  TEXT    NOT NULL,
                product_price REAL    NOT NULL,
                quantity      INTEGER DEFAULT 1,
                image_url     TEXT,
                added_at      TEXT    NOT NULL
            )
        ''')

        # ---- 订单表 ----
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS orders (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id       INTEGER NOT NULL,
                username      TEXT    NOT NULL,
                items_json    TEXT    NOT NULL,
                total_amount  REAL    NOT NULL,
                status        TEXT    DEFAULT 'pending',
                xlsx_path     TEXT,
                created_at    TEXT    NOT NULL
            )
        ''')

        # ---- 浏览历史：已改为纯前端 localStorage 实现（不占服务器存储） ----

        # 常用查询索引，减轻频繁刷新压力
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_consult_status ON consultations(status)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_consult_created ON consultations(created_at)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_subs_created ON subscriptions(created_at)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_orders_user ON orders(user_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_cart_user ON cart_items(user_id)')

        conn.commit()
        conn.close()
        _DB_READY = True
        print("[数据库] 初始化完成 ✓ (WAL 模式)")


# ============================================================
# 静态文件服务 — 让管理后台能访问 CSS/JS
# ============================================================

@app.route('/')
def index_page():
    """网站首页（根路径直达，兼容用户直接访问域名/IP）"""
    return send_from_directory(PROJECT_ROOT, 'index.html')

@app.route('/<path:filename>')
def serve_static(filename):
    """提供项目根目录下的静态文件（CSS/JS等）"""
    # 安全检查：只允许特定扩展名
    safe_exts = ('.css', '.js', '.html', '.jpg', '.jpeg', '.png', '.gif', '.svg', '.ico', '.woff', '.woff2', '.ttf')
    if not any(filename.lower().endswith(ext) for ext in safe_exts):
        return jsonify({'error': 'Forbidden'}), 403
    return send_from_directory(PROJECT_ROOT, filename)


# ============================================================
# 管理后台页面
# ============================================================

@app.route('/admin')
@app.route('/admin/')
def admin_panel():
    """管理后台主页"""
    admin_html = os.path.join(BACKEND_DIR, 'admin.html')
    if os.path.exists(admin_html):
        return send_from_directory(BACKEND_DIR, 'admin.html')
    return '<h1>管理后台文件未找到</h1>', 404


# ============================================================
# API 路由 — 前端调用的接口
# ============================================================

@app.route('/api/health', methods=['GET'])
def health_check():
    """健康检查"""
    return jsonify({
        'status': 'ok',
        'service': '雅檀怡家私城后端',
        'timestamp': datetime.now().isoformat()
    })


@app.route('/api/shutdown', methods=['POST'])
def shutdown_server():
    """★ 优雅关闭 — 仅允许本机调用，防止远程 DoS"""
    if request.remote_addr not in ('127.0.0.1', '::1', 'localhost'):
        return jsonify({'success': False, 'message': 'Forbidden'}), 403
    import os, threading
    def _quit():
        import time; time.sleep(0.5)
        os._exit(0)
    threading.Thread(target=_quit, daemon=True).start()
    return jsonify({'success': True, 'message': '服务正在关闭...'})


@app.route('/api/consult', methods=['POST'])
def submit_consult():
    """
    客户提交免费设计咨询
    前端 JSON: { name, phone, email?, consult_type?, message? }
    """
    try:
        data = request.get_json()

        name = (data.get('name') or '').strip()
        phone = (data.get('phone') or '').strip()

        if not name or not phone:
            return jsonify({'success': False, 'message': '姓名和电话不能为空'}), 400

        if len(name) > 50:
            return jsonify({'success': False, 'message': '姓名过长'}), 400

        if not phone.isdigit() or len(phone) != 11:
            return jsonify({'success': False, 'message': '手机号格式不正确'}), 400

        email = (data.get('email') or '').strip()
        consult_type = (data.get('consult_type') or '').strip()
        message = (data.get('message') or '').strip()
        ip = request.remote_addr or ''

        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO consultations (name, phone, email, consult_type, message, created_at, ip_address)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (name, phone, email, consult_type, message,
              datetime.now().isoformat(), ip))
        conn.commit()
        new_id = cursor.lastrowid

        print(f"[咨询] 新记录 #{new_id}: {name} {phone} ({consult_type})")

        return jsonify({
            'success': True,
            'message': '提交成功！我们的设计师将在24小时内与您联系。',
            'id': new_id
        })

    except Exception as e:
        print(f"[错误] 提交咨询失败: {e}")
        return jsonify({'success': False, 'message': '服务器错误，请稍后重试'}), 500


@app.route('/api/subscribe', methods=['POST'])
def subscribe_newsletter():
    """新闻/资讯订阅"""
    try:
        data = request.get_json()
        email = (data.get('email') or '').strip()

        if not email or '@' not in email:
            return jsonify({'success': False, 'message': '请输入有效的邮箱地址'}), 400

        ip = request.remote_addr or ''

        conn = get_db()
        cursor = conn.cursor()

        cursor.execute('SELECT id FROM subscriptions WHERE email = ?', (email,))
        if cursor.fetchone():

            return jsonify({'success': True, 'message': '您已订阅，无需重复订阅'})

        cursor.execute(
            'INSERT INTO subscriptions (email, created_at, ip_address) VALUES (?, ?, ?)',
            (email, datetime.now().isoformat(), ip))
        conn.commit()

        print(f"[订阅] {email}")
        return jsonify({'success': True, 'message': '订阅成功！感谢您的关注。'})

    except Exception as e:
        print(f"[错误] 订阅失败: {e}")
        return jsonify({'success': False, 'message': '服务器错误'}), 500


# ============================================================
# 管理端 API — 后台数据管理
# ============================================================

@app.route('/api/admin/login', methods=['POST'])
def admin_login():
    """管理后台登录：密码正确则写入 session（含失败锁定防暴力破解）"""
    data = request.get_json() or {}
    pwd = (data.get('password') or '').strip()
    ip = request.remote_addr or 'unknown'
    # ★ 安全加固：同IP连续失败5次锁定10分钟
    lock_left = _check_login_lock(ip)
    if lock_left is not None:
        return jsonify({'success': False, 'message': f'尝试次数过多，请 {lock_left} 秒后再试'}), 429
    if pwd == ADMIN_PASSWORD:
        session['admin_auth'] = True
        session.permanent = False          # 关闭浏览器即失效
        return jsonify({'success': True, 'message': '登录成功'})
    _record_login_fail(ip)
    return jsonify({'success': False, 'message': '密码错误'}), 401


@app.route('/api/admin/logout', methods=['POST'])
def admin_logout():
    """退出登录"""
    session.pop('admin_auth', None)
    return jsonify({'success': True, 'message': '已退出'})


@app.route('/api/admin/check', methods=['GET'])
def admin_check():
    """检查是否已登录（前端加载时调用）"""
    if session.get('admin_auth'):
        return jsonify({'success': True})
    return jsonify({'success': False}), 401


@app.route('/api/stats', methods=['GET'])
@admin_required
def get_stats():
    """
    数据看板统计（一次返回全部看板数字，减少刷新时的请求数）
    """
    try:
        conn = get_db()
        cursor = conn.cursor()
        today = datetime.now().strftime('%Y-%m-%d')

        cursor.execute('SELECT COUNT(*) FROM consultations')
        total_consult = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM consultations WHERE status='pending'")
        pending = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM consultations WHERE created_at LIKE ?", (today + '%',))
        today_new = cursor.fetchone()[0]
        cursor.execute('SELECT COUNT(*) FROM subscriptions')
        total_subs = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM subscriptions WHERE created_at LIKE ?", (today + '%',))
        today_subs = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM consultations WHERE status='completed'")
        completed = cursor.fetchone()[0]
        cursor.execute('SELECT COUNT(*) FROM users')
        total_users = cursor.fetchone()[0]
        cursor.execute('SELECT COUNT(*) FROM orders')
        total_orders = cursor.fetchone()[0]
        cursor.execute('SELECT COUNT(*) FROM products')
        total_products = cursor.fetchone()[0]

        return jsonify({
            'success': True,
            'data': {
                'total_consult': total_consult,
                'pending': pending,
                'today_new': today_new,
                'completed': completed,
                'total_subs': total_subs,
                'today_subs': today_subs,
                'total_users': total_users,
                'total_orders': total_orders,
                'total_products': total_products,
            }
        })
    except Exception as e:
        print(f"[错误] stats: {e}")
        return jsonify({'success': False, 'message': '统计查询失败'}), 500


@app.route('/api/admin/consultations', methods=['GET'])
@admin_required
def list_consultations():
    """
    查看所有咨询记录
    支持参数：?status=pending&page=1&limit=20
    """
    status_filter = request.args.get('status', '')
    # ★ BUG修复：非法 page/limit 不再直接 500，且 limit 设上限防止一次拉全表
    try:
        page = max(1, int(request.args.get('page', 1)))
    except ValueError:
        page = 1
    try:
        limit = min(100, max(1, int(request.args.get('limit', 50))))
    except ValueError:
        limit = 50
    offset = (page - 1) * limit

    conn = get_db()
    cursor = conn.cursor()

    # 条件查询
    if status_filter:
        cursor.execute('SELECT COUNT(*) FROM consultations WHERE status = ?', (status_filter,))
        total = cursor.fetchone()[0]
        cursor.execute(
            'SELECT * FROM consultations WHERE status = ? ORDER BY created_at DESC LIMIT ? OFFSET ?',
            (status_filter, limit, offset))
    else:
        cursor.execute('SELECT COUNT(*) FROM consultations')
        total = cursor.fetchone()[0]
        cursor.execute(
            'SELECT * FROM consultations ORDER BY created_at DESC LIMIT ? OFFSET ?',
            (limit, offset))

    rows = [dict(row) for row in cursor.fetchall()]

    return jsonify({
        'success': True,
        'data': rows,
        'total': total,
        'page': page,
        'limit': limit
    })


@app.route('/api/admin/subscriptions', methods=['GET'])
@admin_required
def list_subscriptions():
    """查看所有订阅记录"""
    try:
        page = max(1, int(request.args.get('page', 1)))
    except ValueError:
        page = 1
    try:
        limit = min(100, max(1, int(request.args.get('limit', 50))))
    except ValueError:
        limit = 50
    offset = (page - 1) * limit

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT COUNT(*) FROM subscriptions')
    total = cursor.fetchone()[0]
    cursor.execute('SELECT * FROM subscriptions ORDER BY created_at DESC LIMIT ? OFFSET ?', (limit, offset))
    rows = [dict(row) for row in cursor.fetchall()]

    return jsonify({
        'success': True,
        'data': rows,
        'total': total,
        'page': page,
        'limit': limit
    })


@app.route('/api/admin/consultations/<int:cid>/status', methods=['PUT'])
@admin_required
def update_consultation_status(cid):
    """更新咨询状态：pending → contacted → completed"""
    data = request.get_json()
    new_status = (data.get('status') or '').strip()

    if new_status not in ('pending', 'contacted', 'completed'):
        return jsonify({'success': False, 'message': '无效状态'}), 400

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('UPDATE consultations SET status = ? WHERE id = ?', (new_status, cid))
    conn.commit()
    affected = cursor.rowcount

    if affected == 0:
        return jsonify({'success': False, 'message': '记录不存在'}), 404

    return jsonify({'success': True, 'message': f'状态已更新为 {new_status}'})


@app.route('/api/admin/consultations/<int:cid>', methods=['DELETE'])
@admin_required
def delete_consultation(cid):
    """删除咨询记录"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM consultations WHERE id = ?', (cid,))
    conn.commit()
    affected = cursor.rowcount

    if affected == 0:
        return jsonify({'success': False, 'message': '记录不存在'}), 404

    return jsonify({'success': True, 'message': '记录已删除'})


@app.route('/api/admin/subscriptions/<int:sid>', methods=['DELETE'])
@admin_required
def delete_subscription(sid):
    """删除订阅记录"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM subscriptions WHERE id = ?', (sid,))
    conn.commit()
    affected = cursor.rowcount

    if affected == 0:
        return jsonify({'success': False, 'message': '记录不存在'}), 404

    return jsonify({'success': True, 'message': '订阅已删除'})


# ============================================================
# ★★★ 用户认证 API ★★★
# ============================================================

def _hash_pw(pw):
    """密码哈希：优先 werkzeug 的 PBKDF2（慢哈希，抗彩虹表）；
       兼容旧版纯 sha256 存储的密码（登录时自动升级）"""
    try:
        from werkzeug.security import generate_password_hash, check_password_hash
        return generate_password_hash(pw, method='pbkdf2:sha256')
    except Exception:
        return 'sha256$' + hashlib.sha256(pw.encode()).hexdigest()

def _check_pw(stored, pw):
    """校验密码；兼容 sha256 旧格式（匹配后返回 'upgrade' 表示需要升级存储）"""
    try:
        from werkzeug.security import check_password_hash
        if stored.startswith(('pbkdf2:', 'scrypt:', 'sha256$')):
            return 'ok' if check_password_hash(stored, pw) else 'no'
    except Exception:
        pass
    # 旧版裸 sha256
    if stored == hashlib.sha256(pw.encode()).hexdigest():
        return 'upgrade'
    return 'no'

@app.route('/api/auth/register', methods=['POST'])
def register():
    """用户注册"""
    import re
    data = request.get_json()
    username = (data.get('username') or '').strip()
    password = (data.get('password') or '').strip()
    phone = (data.get('phone') or '').strip()
    if not username:
        return jsonify({'success': False, 'message': '请填写用户名'}), 400
    if not password or len(password) < 6:
        return jsonify({'success': False, 'message': '密码至少6位'}), 400
    if not re.search(r'[a-zA-Z]', password) or not re.search(r'\d', password):
        return jsonify({'success': False, 'message': '密码必须包含字母和数字的组合'}), 400
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT id FROM users WHERE username=?', (username,))
    if c.fetchone():

        return jsonify({'success': False, 'message': '用户名已存在'}), 400
    token = uuid.uuid4().hex
    c.execute('INSERT INTO users (username,password,phone,created_at,last_login) VALUES (?,?,?,?,?)',
              (username, _hash_pw(password), phone, datetime.now().isoformat(), datetime.now().isoformat()))
    user_id = c.lastrowid
    conn.commit()
    # ★ 安全加固：注册成功即建立服务端会话（不再信任前端传参）
    session['user_id'] = user_id
    # 创建购物车token（用token做客户端身份标识）
    return jsonify({'success': True, 'message': '欢迎加入雅檀怡家私城，请完善个人信息', 'user': {'id': user_id, 'username': username, 'token': token}})

@app.route('/api/auth/login', methods=['POST'])
def login():
    """用户登录（成功后建立服务端 session）"""
    data = request.get_json() or {}
    username = (data.get('username') or '').strip()
    password = (data.get('password') or '').strip()
    if not username or not password:
        return jsonify({'success': False, 'message': '请填写用户名和密码'}), 400
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT * FROM users WHERE username=?', (username,))
    row = c.fetchone()
    if not row:
        return jsonify({'success': False, 'message': '用户名或密码错误'}), 401
    pw_check = _check_pw(row['password'], password)
    if pw_check == 'no':
        return jsonify({'success': False, 'message': '用户名或密码错误'}), 401
    # 旧版 sha256 密码匹配成功 → 自动升级为新哈希存储
    if pw_check == 'upgrade':
        c.execute('UPDATE users SET password=? WHERE id=?', (_hash_pw(password), row['id']))
    c.execute('UPDATE users SET last_login=? WHERE id=?', (datetime.now().isoformat(), row['id']))
    conn.commit()
    # ★ 安全加固：登录成功写入服务端 session
    session['user_id'] = row['id']
    keys = set(row.keys())
    return jsonify({'success': True, 'message': '登录成功！',
        'user': {
            'id': row['id'],
            'username': row['username'],
            'phone': row['phone'] if 'phone' in keys else '',
            'email': row['email'] if 'email' in keys else '',
            'nickname': (row['nickname'] if 'nickname' in keys and row['nickname'] else row['username']),
            'gender': row['gender'] if 'gender' in keys else '',
            'age': row['age'] if 'age' in keys else 0,
            'region': row['region'] if 'region' in keys else '',
            'signature': row['signature'] if 'signature' in keys else '',
            'avatar': row['avatar'] if 'avatar' in keys else '',
        }})

@app.route('/api/auth/logout', methods=['POST'])
def logout():
    """退出登录：清除服务端会话"""
    session.pop('user_id', None)
    return jsonify({'success': True, 'message': '已退出'})


# ============================================================
# ★★★ 忘记密码 API（验证码方式） ★★★
# ============================================================

import random as _random

# 验证码发送限流：同一手机号60秒内只允许发1条
_sms_gate = {}  # {phone: last_send_time}

@app.route('/api/auth/forgot-password', methods=['POST'])
def forgot_password():
    """发送验证码到手机（演示模式：验证码仅在服务器日志打印；
       生产环境请接入短信服务商）"""
    data = request.get_json()
    phone = (data.get('phone') or '').strip()
    if not phone or len(phone) != 11 or not phone.isdigit():
        return jsonify({'success': False, 'message': '请输入正确的11位手机号'}), 400
    # ★ 安全加固：限流，防止验证码轰炸
    last = _sms_gate.get(phone)
    if last and (datetime.now() - last).seconds < 60:
        return jsonify({'success': False, 'message': '发送过于频繁，请60秒后再试'}), 429
    conn = get_db(); c = conn.cursor()
    c.execute('SELECT id,username FROM users WHERE phone=?', (phone,))
    user = c.fetchone()
    if not user:

        return jsonify({'success': False, 'message': '该手机号未注册'}), 404
    # 生成6位验证码，有效期5分钟
    code = str(_random.randint(100000, 999999))
    expires = (datetime.now() + timedelta(minutes=5)).isoformat()
    c.execute('INSERT INTO verification_codes (phone,code,purpose,expires_at,created_at) VALUES (?,?,?,?,?)',
              (phone, code, 'reset_password', expires, datetime.now().isoformat()))
    conn.commit()
    _sms_gate[phone] = datetime.now()
    print(f'[验证码] {phone} → {code}（有效至 {expires}）')
    # ★ 安全加固：不再把验证码明文返回前端（生产必须接入短信；演示时看后端日志）
    return jsonify({'success': True, 'message': f'验证码已发送到 {phone}'})

@app.route('/api/auth/verify-code', methods=['POST'])
def verify_code():
    """验证验证码是否正确"""
    data = request.get_json()
    phone = (data.get('phone') or '').strip()
    code = (data.get('code') or '').strip()
    if not phone or not code:
        return jsonify({'success': False, 'message': '参数不完整'}), 400
    conn = get_db(); c = conn.cursor()
    c.execute('''SELECT * FROM verification_codes WHERE phone=? AND code=? AND purpose='reset_password'
                 AND used=0 AND expires_at > ? ORDER BY created_at DESC LIMIT 1''',
              (phone, code, datetime.now().isoformat()))
    row = c.fetchone()
    if not row:

        return jsonify({'success': False, 'message': '验证码错误或已过期'}), 400
    c.execute('UPDATE verification_codes SET used=1 WHERE id=?', (row['id'],))
    conn.commit()
    return jsonify({'success': True, 'message': '验证通过', 'phone': phone})

@app.route('/api/auth/reset-password', methods=['POST'])
def reset_password():
    """重置密码"""
    data = request.get_json()
    phone = (data.get('phone') or '').strip()
    code = (data.get('code') or '').strip()
    password = (data.get('password') or '').strip()
    if not phone or not code or not password or len(password) < 6:
        return jsonify({'success': False, 'message': '密码至少6位'}), 400
    conn = get_db(); c = conn.cursor()
    # 再次验证验证码
    c.execute('''SELECT * FROM verification_codes WHERE phone=? AND code=? AND purpose='reset_password'
                 AND used=1 AND expires_at > ? ORDER BY created_at DESC LIMIT 1''',
              (phone, code, datetime.now().isoformat()))
    if not c.fetchone():

        return jsonify({'success': False, 'message': '验证码无效'}), 400
    c.execute('UPDATE users SET password=? WHERE phone=?', (_hash_pw(password), phone))
    conn.commit()
    return jsonify({'success': True, 'message': '密码重置成功！请重新登录。'})


# ============================================================
# ★★★ 个人资料 API ★★★
# ============================================================

@app.route('/api/user/profile', methods=['GET'])
@require_login
def get_profile():
    """获取用户个人资料（身份取自服务端 session，杜绝越权）"""
    user_id = current_user_id()
    conn = get_db(); c = conn.cursor()
    c.execute('SELECT id,username,phone,email,nickname,gender,age,region,signature,avatar,created_at,last_login FROM users WHERE id=?', (user_id,))
    row = c.fetchone()

    if not row:
        return jsonify({'success': False, 'message': '用户不存在'}), 404
    return jsonify({'success': True, 'user': dict(row)})

@app.route('/api/user/profile', methods=['PUT'])
@require_login
def update_profile():
    """更新个人资料（只允许修改自己的资料）"""
    data = request.get_json()
    user_id = current_user_id()
    fields = ['nickname','gender','age','region','signature','avatar','phone','email']
    updates = {}
    for f in fields:
        if f in data:
            updates[f] = str(data[f]).strip()
    if not updates:
        return jsonify({'success': False, 'message': '没有要更新的字段'}), 400
    conn = get_db(); c = conn.cursor()
    set_clause = ', '.join(f'{k}=?' for k in updates)
    vals = list(updates.values()) + [user_id]
    c.execute(f'UPDATE users SET {set_clause} WHERE id=?', vals)
    conn.commit()
    return jsonify({'success': True, 'message': '资料更新成功！'})

@app.route('/api/user/account', methods=['DELETE'])
@require_login
def delete_account():
    """注销账户（只允许注销自己的账户）"""
    user_id = current_user_id()
    conn = get_db(); c = conn.cursor()
    c.execute('SELECT username FROM users WHERE id=?', (user_id,))
    row = c.fetchone()
    if not row:

        return jsonify({'success': False, 'message': '用户不存在'}), 404
    # 清除个人信息（保留ID和订单关联）
    c.execute('''UPDATE users SET username=?, password='', phone='', email='', 
                 nickname='已注销', gender='', age=0, region='', signature='该用户已注销', avatar='' 
                 WHERE id=?''', ('deleted_' + str(user_id), user_id))
    c.execute('DELETE FROM cart_items WHERE user_id=?', (user_id,))
    conn.commit()
    session.pop('user_id', None)
    return jsonify({'success': True, 'message': '账户已注销。感谢您的使用。'})


# ============================================================
# ★★★ 购物车 API ★★★
# ============================================================

@app.route('/api/cart', methods=['GET'])
@require_login
def get_cart():
    """获取当前登录用户的购物车"""
    user_id = current_user_id()
    conn = get_db(); c = conn.cursor()
    c.execute('SELECT * FROM cart_items WHERE user_id=? ORDER BY added_at DESC', (user_id,))
    items = [dict(r) for r in c.fetchall()]
    total = sum(i['product_price'] * i['quantity'] for i in items)

    return jsonify({'success': True, 'items': items, 'count': len(items), 'total': total})

@app.route('/api/cart', methods=['POST'])
@require_login
def add_to_cart():
    """添加商品到当前登录用户的购物车"""
    data = request.get_json()
    user_id = current_user_id()
    product_name = data.get('product_name')
    product_price = data.get('product_price')
    try:
        quantity = max(1, min(99, int(data.get('quantity', 1))))
    except (TypeError, ValueError):
        quantity = 1
    image_url = data.get('image_url', '')
    if not product_name:
        return jsonify({'success': False, 'message': '参数不完整'}), 400
    try:
        product_price = float(product_price)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'message': '价格无效'}), 400
    conn = get_db(); c = conn.cursor()
    # 检查是否已有同产品
    c.execute('SELECT id,quantity FROM cart_items WHERE user_id=? AND product_name=?',
              (user_id, product_name))
    existing = c.fetchone()
    if existing:
        c.execute('UPDATE cart_items SET quantity=quantity+?, added_at=? WHERE id=?',
                  (quantity, datetime.now().isoformat(), existing['id']))
    else:
        c.execute('INSERT INTO cart_items (user_id,product_name,product_price,quantity,image_url,added_at) VALUES (?,?,?,?,?,?)',
                  (user_id, product_name, product_price, quantity, image_url, datetime.now().isoformat()))
    conn.commit()
    return jsonify({'success': True, 'message': '已加入购物车'})

@app.route('/api/cart/<int:cid>', methods=['PUT'])
@require_login
def update_cart_item(cid):
    """更新购物车数量（★ 只允许操作自己的购物车项）"""
    data = request.get_json()
    user_id = current_user_id()
    try:
        qty = int(data.get('quantity', 1))
    except (TypeError, ValueError):
        return jsonify({'success': False, 'message': '数量无效'}), 400
    qty = max(1, min(99, qty))
    conn = get_db(); c = conn.cursor()
    c.execute('UPDATE cart_items SET quantity=? WHERE id=? AND user_id=?', (qty, cid, user_id))
    conn.commit()
    if c.rowcount == 0:
        return jsonify({'success': False, 'message': '购物车项不存在'}), 404
    return jsonify({'success': True})

@app.route('/api/cart/<int:cid>', methods=['DELETE'])
@require_login
def delete_cart_item(cid):
    """删除购物车项（★ 只允许删除自己的购物车项）"""
    user_id = current_user_id()
    conn = get_db(); c = conn.cursor()
    c.execute('DELETE FROM cart_items WHERE id=? AND user_id=?', (cid, user_id))
    conn.commit()
    if c.rowcount == 0:
        return jsonify({'success': False, 'message': '购物车项不存在'}), 404
    return jsonify({'success': True, 'message': '已移除'})


# ============================================================
# ★★★ 订单 API（生成xlsx） ★★★
# ============================================================

ORDERS_DIR = os.path.join(BACKEND_DIR, '订单文件')
os.makedirs(ORDERS_DIR, exist_ok=True)

def _generate_order_xlsx(order_id, username, items, total):
    """生成订单xlsx文件（只有客服/管理能看）"""
    try:
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active
        ws.title = f'订单_{order_id}'
        ws['A1'] = '订单编号'; ws['B1'] = order_id
        ws['A2'] = '客户姓名'; ws['B2'] = username
        ws['A3'] = '下单时间'; ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A5'] = '产品名称'; ws['B5'] = '单价'; ws['C5'] = '数量'; ws['D5'] = '小计'
        row = 6
        for item in items:
            ws.cell(row, 1, item.get('product_name',''))
            ws.cell(row, 2, item.get('product_price',0))
            ws.cell(row, 3, item.get('quantity',1))
            ws.cell(row, 4, item.get('product_price',0) * item.get('quantity',1))
            row += 1
        ws.cell(row+1, 1, '合计'); ws.cell(row+1, 4, total)
        # ★ 安全加固：文件名只由订单号组成（原拼接 username 可被路径穿越利用）
        filename = f'order_{order_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        filepath = os.path.join(ORDERS_DIR, filename)
        wb.save(filepath)
        return filepath
    except Exception as e:
        print(f'[xlsx] 生成失败: {e}')
        return None

@app.route('/api/orders', methods=['POST'])
@require_login
def create_order():
    """下单：生成订单记录 + xlsx文件（当前登录用户购物车）"""
    data = request.get_json()
    user_id = current_user_id()
    username = data.get('username', '未知用户')
    # ★ 安全加固：xlsx 文件名不再拼接 username（防路径穿越），只由订单号组成
    conn = get_db(); c = conn.cursor()
    c.execute('SELECT * FROM cart_items WHERE user_id=?', (user_id,))
    cart_items = [dict(r) for r in c.fetchall()]
    if not cart_items:

        return jsonify({'success': False, 'message': '购物车为空'}), 400
    items_data = [{'product_name': i['product_name'], 'product_price': i['product_price'],
                    'quantity': i['quantity'], 'image_url': i.get('image_url','')} for i in cart_items]
    total = sum(i['product_price'] * i['quantity'] for i in cart_items)
    # 写订单记录
    c.execute('INSERT INTO orders (user_id,username,items_json,total_amount,created_at) VALUES (?,?,?,?,?)',
              (user_id, username, json.dumps(items_data, ensure_ascii=False), total, datetime.now().isoformat()))
    order_id = c.lastrowid
    # 生成xlsx
    xlsx_path = _generate_order_xlsx(order_id, username, items_data, total)
    if xlsx_path:
        c.execute('UPDATE orders SET xlsx_path=? WHERE id=?', (xlsx_path, order_id))
    # 清空购物车
    c.execute('DELETE FROM cart_items WHERE user_id=?', (user_id,))
    conn.commit()
    return jsonify({'success': True, 'message': f'下单成功！共{len(cart_items)}件，¥{total}',
                    'order_id': order_id, 'total': total, 'item_count': len(cart_items)})

@app.route('/api/orders', methods=['GET'])
@require_login
def get_orders():
    """获取当前登录用户的订单（★ 剔除 xlsx_path，不泄露服务器路径）"""
    user_id = current_user_id()
    conn = get_db(); c = conn.cursor()
    c.execute('SELECT id,user_id,username,items_json,total_amount,status,created_at FROM orders WHERE user_id=? ORDER BY created_at DESC', (user_id,))
    orders = [dict(r) for r in c.fetchall()]

    return jsonify({'success': True, 'orders': orders})


# ============================================================
# ★★★ 产品搜索 API ★★★
# ============================================================

@app.route('/api/products/search', methods=['GET'])
def search_products():
    """产品搜索 + 排序（前台公开接口：只返回上架 is_active=1 的产品）"""
    q = request.args.get('q', '')
    category = request.args.get('category', '')
    sort = request.args.get('sort', '')  # price_asc/price_desc/rating
    conn = get_db(); c = conn.cursor()
    # ★ 老库兼容：无 is_active 列时不加过滤
    cols = [r[1] for r in c.execute('PRAGMA table_info(products)').fetchall()]
    has_active = 'is_active' in cols
    sql = 'SELECT * FROM products WHERE 1=1'
    params = []
    if has_active:
        sql += ' AND is_active=1'
    if q:
        sql += ' AND name LIKE ?'
        params.append(f'%{q}%')
    if category:
        sql += ' AND category=?'
        params.append(category)
    if sort == 'price_asc':
        sql += ' ORDER BY price ASC'
    elif sort == 'price_desc':
        sql += ' ORDER BY price DESC'
    elif sort == 'rating':
        sql += ' ORDER BY rating DESC'
    else:
        sql += ' ORDER BY id DESC'
    c.execute(sql, params)
    rows = [dict(r) for r in c.fetchall()]

    return jsonify({'success': True, 'products': rows, 'total': len(rows)})


# ============================================================
# ★★★ 管理端：用户 & 订单管理 ★★★
# ============================================================

@app.route('/api/admin/users', methods=['GET'])
@admin_required
def admin_users():
    """管理端：查看所有用户"""
    conn = get_db(); c = conn.cursor()
    c.execute('''SELECT id,username,phone,email,nickname,gender,age,region,signature,avatar,created_at,last_login
                 FROM users ORDER BY created_at DESC''')
    rows = [dict(r) for r in c.fetchall()]
    return jsonify({'success': True, 'users': rows, 'total': len(rows)})

@app.route('/api/admin/orders', methods=['GET'])
@admin_required
def admin_orders():
    """管理端：查看所有订单（xlsx_path 只返回文件名，不泄露服务器绝对路径）"""
    conn = get_db(); c = conn.cursor()
    c.execute('SELECT id,user_id,username,items_json,total_amount,status,created_at,xlsx_path FROM orders ORDER BY created_at DESC')
    rows = [dict(r) for r in c.fetchall()]
    for r in rows:
        if r.get('xlsx_path'):
            r['xlsx_path'] = os.path.basename(r['xlsx_path'])
    return jsonify({'success': True, 'orders': rows, 'total': len(rows)})


@app.route('/api/admin/orders/<int:oid>/status', methods=['PUT'])
@admin_required
def update_order_status(oid):
    """
    管理端：手动更新订单状态
    合法状态：pending / confirmed / shipped / completed
    """
    data = request.get_json() or {}
    new_status = (data.get('status') or '').strip()
    if new_status not in ('pending', 'confirmed', 'shipped', 'completed'):
        return jsonify({'success': False, 'message': '无效订单状态'}), 400
    conn = get_db(); c = conn.cursor()
    c.execute('UPDATE orders SET status=? WHERE id=?', (new_status, oid))
    conn.commit()
    if c.rowcount == 0:
        return jsonify({'success': False, 'message': '订单不存在'}), 404
    return jsonify({'success': True, 'message': f'订单状态已更新为 {new_status}', 'status': new_status})


@app.route('/api/admin/orders/<int:oid>/xlsx', methods=['GET'])
@admin_required
def download_order_xlsx(oid):
    """
    下载订单 xlsx（管理端）
    业务规则：客服下载订单文件 = 已查看并确认，自动将 pending → confirmed
    """
    conn = get_db(); c = conn.cursor()
    c.execute('SELECT xlsx_path, status, username, items_json, total_amount FROM orders WHERE id=?', (oid,))
    row = c.fetchone()
    if not row:
        return jsonify({'success': False, 'message': '订单不存在'}), 404

    # ★ 兜底：xlsx 文件丢失/从未生成时，用订单表内的 items_json 快照自动重建
    #   设计原则：数据库 items_json 是唯一事实源，xlsx 只是派生物，丢了可随时再生
    xlsx_path = row['xlsx_path']
    if not xlsx_path or not os.path.exists(xlsx_path):
        try:
            items = json.loads(row['items_json']) if row['items_json'] else []
            xlsx_path = _generate_order_xlsx(oid, row['username'], items, row['total_amount'])
            if not xlsx_path:
                return jsonify({'success': False, 'message': '文件重建失败'}), 500
            c.execute('UPDATE orders SET xlsx_path=? WHERE id=?', (xlsx_path, oid))
            conn.commit()
            print(f"[xlsx] 订单 #{oid} 文件丢失，已根据 items_json 自动重建: {os.path.basename(xlsx_path)}")
        except Exception as e:
            print(f'[xlsx] 订单 #{oid} 文件重建失败: {e}')
            return jsonify({'success': False, 'message': '文件重建失败'}), 500

    # ★ BUG修复：下载即确认 —— 只有待确认订单才自动推进状态
    if row['status'] == 'pending':
        c.execute("UPDATE orders SET status='confirmed' WHERE id=?", (oid,))
        conn.commit()
        print(f"[订单] #{oid} 下载后自动确认为 confirmed")

    return send_file(
        xlsx_path,
        as_attachment=True,
        download_name=os.path.basename(xlsx_path)
    )


# ============================================================
# ★★★ 管理端：产品管理 ★★★
# ============================================================

VALID_CATEGORIES = ('sofa', 'bed', 'table', 'storage')
ALLOWED_IMG_EXT = ('.jpg', '.jpeg', '.png', '.gif', '.webp')

def _ensure_products_cols(conn):
    """老库迁移：products 表补充 is_active（上架状态）字段"""
    c = conn.cursor()
    cols = [r[1] for r in c.execute('PRAGMA table_info(products)').fetchall()]
    if 'is_active' not in cols:
        c.execute('ALTER TABLE products ADD COLUMN is_active INTEGER DEFAULT 1')
        conn.commit()

@app.route('/api/admin/products', methods=['GET'])
@admin_required
def admin_products():
    """管理端：产品列表（分页 + 关键词 + 分类过滤）"""
    try:
        page = max(1, int(request.args.get('page', 1) or 1))
    except ValueError:
        page = 1
    try:
        limit = min(100, max(1, int(request.args.get('limit', 20) or 20)))
    except ValueError:
        limit = 20
    offset = (page - 1) * limit
    q = (request.args.get('q') or '').strip()
    category = (request.args.get('category') or '').strip()

    conn = get_db(); c = conn.cursor()
    _ensure_products_cols(conn)
    where, params = '1=1', []
    if q:
        where += ' AND name LIKE ?'
        params.append(f'%{q}%')
    if category:
        where += ' AND category=?'
        params.append(category)

    c.execute(f'SELECT COUNT(*) FROM products WHERE {where}', params)
    total = c.fetchone()[0]
    c.execute(f'SELECT * FROM products WHERE {where} ORDER BY id DESC LIMIT ? OFFSET ?',
              params + [limit, offset])
    rows = [dict(r) for r in c.fetchall()]
    return jsonify({'success': True, 'products': rows, 'total': total, 'page': page, 'limit': limit})


@app.route('/api/admin/products/<int:pid>', methods=['GET'])
@admin_required
def admin_product_detail(pid):
    """管理端：单个产品详情"""
    conn = get_db(); c = conn.cursor()
    c.execute('SELECT * FROM products WHERE id=?', (pid,))
    row = c.fetchone()
    if not row:
        return jsonify({'success': False, 'message': '产品不存在'}), 404
    return jsonify({'success': True, 'product': dict(row)})


@app.route('/api/admin/products', methods=['POST'])
@admin_required
def admin_create_product():
    """管理端：新增产品"""
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    category = (data.get('category') or '').strip()
    if not name:
        return jsonify({'success': False, 'message': '产品名称不能为空'}), 400
    if category not in VALID_CATEGORIES:
        return jsonify({'success': False, 'message': '无效的产品分类'}), 400
    try:
        price = float(data.get('price', 0) or 0)
        original_price = float(data.get('original_price', 0) or 0) or None
        rating = float(data.get('rating', 4.5) or 4.5)
        review_count = int(data.get('review_count', 0) or 0)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'message': '价格/评分/评论数格式不正确'}), 400
    if price <= 0:
        return jsonify({'success': False, 'message': '价格必须大于 0'}), 400
    if not (0 <= rating <= 5):
        return jsonify({'success': False, 'message': '评分必须在 0~5 之间'}), 400

    description = (data.get('description') or '').strip()
    image_url = (data.get('image_url') or '').strip()
    is_hot = 1 if data.get('is_hot') else 0
    is_new = 1 if data.get('is_new') else 0
    is_active = 1 if data.get('is_active', 1) else 0

    conn = get_db(); c = conn.cursor()
    _ensure_products_cols(conn)
    c.execute('''INSERT INTO products (name, category, price, original_price, description,
                 image_url, rating, review_count, is_hot, is_new, is_active, created_at)
                 VALUES (?,?,?,?,?,?,?,?,?,?,?,?)''',
              (name, category, price, original_price, description, image_url,
               rating, review_count, is_hot, is_new, is_active, datetime.now().isoformat()))
    conn.commit()
    pid = c.lastrowid
    print(f"[产品] 新增 #{pid}: {name} ¥{price} ({category})")
    return jsonify({'success': True, 'message': '产品已新增', 'id': pid})


@app.route('/api/admin/products/<int:pid>', methods=['PUT'])
@admin_required
def admin_update_product(pid):
    """管理端：编辑产品（部分更新：只修改传入的字段，其余保留）"""
    conn = get_db(); c = conn.cursor()
    c.execute('SELECT * FROM products WHERE id=?', (pid,))
    row = c.fetchone()
    if not row:
        return jsonify({'success': False, 'message': '产品不存在'}), 404
    old = dict(row)

    data = request.get_json() or {}
    name = (data.get('name') if 'name' in data else old['name'] or '').strip()
    category = (data.get('category') if 'category' in data else old['category'] or '').strip()
    if not name:
        return jsonify({'success': False, 'message': '产品名称不能为空'}), 400
    if category not in VALID_CATEGORIES:
        return jsonify({'success': False, 'message': '无效的产品分类'}), 400

    price = old['price']; original_price = old['original_price']
    rating = old['rating'] if old['rating'] is not None else 4.5
    review_count = old['review_count'] or 0
    try:
        if 'price' in data:
            price = float(data.get('price', 0) or 0)
        if 'original_price' in data:
            original_price = float(data.get('original_price', 0) or 0) or None
        if 'rating' in data:
            rating = float(data.get('rating', 4.5) or 4.5)
        if 'review_count' in data:
            review_count = int(data.get('review_count', 0) or 0)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'message': '价格/评分/评论数格式不正确'}), 400
    if price <= 0:
        return jsonify({'success': False, 'message': '价格必须大于 0'}), 400
    if not (0 <= rating <= 5):
        return jsonify({'success': False, 'message': '评分必须在 0~5 之间'}), 400

    description = (data.get('description') if 'description' in data else old['description'] or '').strip()
    image_url = (data.get('image_url') if 'image_url' in data else old['image_url'] or '').strip()
    is_hot = 1 if data.get('is_hot', old['is_hot']) else 0
    is_new = 1 if data.get('is_new', old['is_new']) else 0
    is_active = 1 if data.get('is_active', old['is_active']) else 0

    c.execute('''UPDATE products SET name=?, category=?, price=?, original_price=?, description=?,
                 image_url=?, rating=?, review_count=?, is_hot=?, is_new=?, is_active=?
                 WHERE id=?''',
              (name, category, price, original_price, description, image_url,
               rating, review_count, is_hot, is_new, is_active, pid))
    conn.commit()
    print(f"[产品] 更新 #{pid}: {name}")
    return jsonify({'success': True, 'message': '产品已更新'})


@app.route('/api/admin/products/<int:pid>', methods=['DELETE'])
@admin_required
def admin_delete_product(pid):
    """管理端：删除产品"""
    conn = get_db(); c = conn.cursor()
    c.execute('DELETE FROM products WHERE id=?', (pid,))
    conn.commit()
    if c.rowcount == 0:
        return jsonify({'success': False, 'message': '产品不存在'}), 404
    print(f"[产品] 删除 #{pid}")
    return jsonify({'success': True, 'message': '产品已删除'})


@app.route('/api/admin/upload', methods=['POST'])
@admin_required
def admin_upload_image():
    """管理端：产品图片上传 → 保存到 images/uploads/，返回相对 URL"""
    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'success': False, 'message': '未选择文件'}), 400
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ALLOWED_IMG_EXT:
        return jsonify({'success': False, 'message': '仅支持 JPG/PNG/GIF/WebP 图片'}), 400
    upload_dir = os.path.join(PROJECT_ROOT, 'images', 'uploads')
    os.makedirs(upload_dir, exist_ok=True)
    fname = f'p_{datetime.now().strftime("%Y%m%d%H%M%S")}_{uuid.uuid4().hex[:8]}{ext}'
    f.save(os.path.join(upload_dir, fname))
    url = f'images/uploads/{fname}'
    print(f"[产品] 图片上传: {url}")
    return jsonify({'success': True, 'message': '上传成功', 'url': url})


# ============================================================
# 启动入口
# ============================================================

if __name__ == '__main__':
    init_db()

    print("=" * 60)
    print("  雅檀怡家私城 - 后端服务")
    print(f"  项目目录: {PROJECT_ROOT}")
    print("  前端首页: http://localhost:5000/index.html")
    print("  管理后台: http://localhost:5000/admin")
    print("  健康检查: http://localhost:5000/api/health")
    print("=" * 60)

    # ★ 用 waitress 优先（生产级多线程）；没有则回退 Werkzeug
    # ★ 关键解决「频繁刷新后连接超时、无法重连」：
    #   1) SQLite WAL + busy_timeout（上文）
    #   2) 请求级连接 teardown（上文）
    #   3) 多线程服务器 channel/connection 超时，避免半开连接占死线程
    try:
        from waitress import serve
        print("  [服务器] waitress (生产模式)")
        serve(
            app,
            host='0.0.0.0',
            port=5000,
            threads=32,              # ★ 8→32：刷新风暴时线程不再被打满
            channel_timeout=20,
            connection_limit=200,    # ★ 100→200：并发连接上限翻倍
            cleanup_interval=10,
        )
    except ImportError:
        print("  [服务器] werkzeug (开发模式) — 建议: pip install waitress")
        app.run(host='0.0.0.0', port=5000, debug=True, threaded=True, use_reloader=False)
